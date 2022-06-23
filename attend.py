
from selenium import webdriver
import time
from openpyxl import Workbook,load_workbook

file_excel="./accountlogin.xlsx"

Excel_worksheet = "Sheet1"
wb=load_workbook(file_excel)
ws = wb[Excel_worksheet]
row_num=ws.max_row
col_num=ws.max_column
driver = webdriver.Chrome()
driver.get("https://fap.fpt.edu.vn/Default.aspx")
driver.set_window_size(1920, 1080)
driver.find_element_by_xpath('/html/body/div/div[2]/div/form/table/tbody/tr[1]/td/div/div/div/div[2]/div/fieldset/div/center/div/div[1]/select').click();
time.sleep(1)
driver.find_element_by_xpath('/html/body/div/div[2]/div/form/table/tbody/tr[1]/td/div/div/div/div[2]/div/fieldset/div/center/div/div[1]/select/option[3]').click();
time.sleep(1)
driver.find_element_by_xpath('/html/body/div/div[2]/div/form/table/tbody/tr[1]/td/div/div/div/div[2]/div/fieldset/div/center/div/div[2]/div/div/div').click();

window_before = driver.window_handles[0]
window_after = driver.window_handles[1]
driver.switch_to.window(window_after)



driver.find_element_by_xpath('/html/body/div[1]/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div[1]/div/form/span/section/div/div/div[1]/div/div[1]/div/div[1]/input').send_keys(ws.cell(1,1).value)
driver.find_element_by_xpath('/html/body/div[1]/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div[2]/div/div[1]/div/div/button/span').click()
time.sleep(2)
driver.find_element_by_xpath('/html/body/div[1]/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div[1]/div/form/span/section/div/div/div[1]/div[1]/div/div/div/div/div[1]/div/div[1]/input').send_keys(ws.cell(2,1).value)
driver.find_element_by_xpath('/html/body/div[1]/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div[2]/div/div[1]/div/div/button/span').click()
driver.switch_to.window(window_before)


time.sleep(6)
driver.get("https://fap.fpt.edu.vn/Teacher.aspx")

time.sleep(1.5)
driver.find_element_by_xpath('/html/body/div/div[2]/div/form/table/tbody/tr[1]/td/div/table/tbody/tr[1]/td[2]/ul/li[1]/a').click()

driver.get('https://fap.fpt.edu.vn/Attendance/TakeAttendance.aspx')


element = driver.find_element_by_xpath('//a[text()="Take"]')
driver.find_element_by_xpath('//a[text()="Take"]').click()
driver.get(element.get_attribute('href'))
driver.execute_script('var list =document.querySelectorAll("input[value=rdPresent]"); list.forEach((i) =>{i.click()})')



time.sleep(1)

driver.find_element_by_xpath('//input[@type="submit"]').click()
driver.close();
    


